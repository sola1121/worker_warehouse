import datetime
import logging

import pytz
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse
from django.core.paginator import Paginator

from account.models import History
from warehouse import models, forms

# Create your views here.

OK = "OK"
TZ = pytz.timezone("Asia/Shanghai")
PER_PAGE = 10

### 供应商相关 ###

@login_required(login_url="/")
def supplier(request, page=1):
    """供应商页面显示与查询"""
    supplier_id = request.GET.get("supplierId", '')
    supplier_name = request.GET.get("supplierName", '')
    suppliers = models.Supplier.objects.filter(is_deleted=False, 
                                               supplier_id__icontains=supplier_id, 
                                               supplier_name__icontains=supplier_name)
    paginator = Paginator(suppliers.order_by("-id"), PER_PAGE)
    if int(page) not in list(paginator.page_range):
        return redirect("/warehouse/supplier/")
    pagetor = paginator.get_page(page)
    return render(request, 
                  "app_warehouse/supplier.html", 
                  {"active_navbar": "supplier", 
                   "pagetor": pagetor, 
                   "supplier_id": supplier_id, 
                   "supplier_name": supplier_name})


@login_required(login_url="/")
def supplier_modify(request):
    """供应商添加和修改"""
    supplier_id = request.GET.get("supplierId", '')
    form = forms.SupplierForm()
    if supplier_id:
        supplier = get_object_or_404(models.Supplier, supplier_id=supplier_id)
        form = forms.SupplierForm(instance=supplier)

    if request.method == "POST":
        action = History.CREATE   # 记录操作动作, 初始默认是创建
        origin_supplier_id = request.POST.get("origin_supplier_id", '')
        if origin_supplier_id:
            try:
                action = History.MODIFY
                change_supplier_id = request.POST.get("supplier_id")
                if origin_supplier_id != change_supplier_id:
                    is_exsit_supplier = models.Supplier.objects.filter(supplier_id=change_supplier_id).exists()   # 取保更改的唯一性
                    if is_exsit_supplier:
                        raise AssertionError("supplier objects with %s has already existed." % change_supplier_id)
                origin_supplier = models.Supplier.objects.get(supplier_id=origin_supplier_id)
            except AssertionError:
                return JsonResponse({"back_msg": "%s 供应商编号已存在."%change_supplier_id})
            except:
                return JsonResponse({"back_msg": "源数据取出失败."})
        pk_id = origin_supplier.id   # 保留对象的pk, 以免关联的表出错
        origin_supplier.delete()
        form = forms.SupplierForm(request.POST)
        if form.is_valid():
            new_supplier = form.save(commit=False)
            new_supplier.id = pk_id
            new_supplier.save()
            # NOTE:　向history中存入记录
            new_history = History.set_record(cur_user=request.user, model=new_supplier, act=action)
            new_history.save()
            return JsonResponse({"back_msg": OK})
        else:
            origin_supplier.save()   # 恢复原始数据的删除
            full_msg = str()
            for value in form.errors.values():
                for msg in value: 
                    full_msg += msg + "\n" 
            return JsonResponse({"back_msg": full_msg})

    return render(request, 
                  "app_warehouse/supplier_change.html", 
                  {"head_title": "编辑供应商",
                   "active_navbar": "supplier",
                   "form": form})


@login_required(login_url="/")
def supplier_delete(request):
    """供应商删除"""
    supplier_id = request.POST.get("supplierId", '')
    try:
        supplier = models.Supplier.objects.get(supplier_id=supplier_id)
    except:
        return JsonResponse({"back_msg": "数据库出错, 未能正确删除内容."})
    supplier.is_deleted = True
    supplier.supplier_id = str(supplier.supplier_id) + "(DELETED %s)"%datetime.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S %f")   # 避免删除后对添加或修改会有唯一性冲突
    supplier.save()
    # NOTE: 向history中存入记录
    new_history = History.set_record(cur_user=request.user, model=supplier, act=History.DELETE)
    new_history.save()
    return JsonResponse({"back_msg": OK})


@login_required(login_url="/")
def supplier_download(request):
    """下载查询结果"""
    supplier_id = request.GET.get("supplierId", '')
    supplier_name = request.GET.get("supplierName", '')
    suppliers = models.Supplier.objects.filter(is_deleted=False, 
                                               supplier_id__icontains=supplier_id, 
                                               supplier_name__icontains=supplier_name)
    filename = "供应商 - 搜索编号_%s, 搜索名称_%s.xlsx" % (supplier_id, supplier_name) if any((supplier_id, supplier_name)) else "供应商.xlsx"
    file_wb = Workbook()
    file_sheet = file_wb.get_sheet_by_name(file_wb.sheetnames[0])
    file_sheet.title = "供应商"
    file_sheet.append(["供应商编号", "供应商名称", "备注"])
    for info in suppliers.values():
        file_sheet.append([info["supplier_id"], info["supplier_name"], info["remark"].replace('\n', '').replace('\r', '')])
    response = HttpResponse(save_virtual_workbook(file_wb))
    response["Content-Type"] = "application/vnd.ms-excel"
    response["Content-Disposition"] = "attachment;filename=\"%s\"" % filename
    return response


### 物品种类相关 ###

@login_required(login_url="/")
def classification(request, page=1):
    """物品分类的查询与显示"""
    class_name = request.GET.get("className", '')
    classifications = models.Classification.objects.filter(is_deleted=False, 
                                                           class_name__icontains=class_name)
    paginator = Paginator(classifications.order_by("-id"), PER_PAGE)
    if int(page) not in list(paginator.page_range):
        return redirect("/warehouse/classification/")
    pagetor = paginator.get_page(page)
    return render(request, 
                  "app_warehouse/classification.html", 
                  {"active_navbar": "classification", 
                   "pagetor": pagetor, 
                   "class_name": class_name})


@login_required(login_url="/")
def classification_modify(request):
    """添加和更改相应的物品分类"""
    class_name = request.GET.get("className", '')
    form = forms.ClassificationForm()
    if class_name:
        classification = get_object_or_404(models.Classification, class_name=class_name)
        form = forms.ClassificationForm(instance=classification)

    if request.method == "POST":
        action = History.CREATE   # 记录操作动作, 初始默认是创建
        origin_class_name = request.POST.get("origin_class_name", '')
        if origin_class_name:
            try:
                action = History.MODIFY
                change_class_name = request.POST.get("class_name")
                if origin_class_name != change_class_name:
                    is_exist_classification = models.Classification.objects.filter(class_name=change_class_name).exists()
                    if is_exist_classification:
                        raise AssertionError("classification objects with %s has aready existed." % change_class_name)
                origin_classification = models.Classification.objects.get(class_name=origin_class_name)
            except AssertionError:
                return JsonResponse({"back_msg": "%s 种类名已存在."%change_class_name})
            except:
                return JsonResponse({"back_msg": "源数据取出失败."})
        pk_id = origin_classification.id   # 保留对象的pk, 以免关联的表出错
        origin_classification.delete()
        form = forms.ClassificationForm(request.POST)
        if form.is_valid():
            new_classification = form.save(commit=False)
            new_classification.id = pk_id
            new_classification.save()
            # NOTE: 向history中存入记录
            new_history = History.set_record(cur_user=request.user, model=new_classification, act=action)
            new_history.save()
            return JsonResponse({"back_msg": OK})
        else:
            origin_classification.save()   # 恢复原始数据删除
            full_msg = str()
            for value in form.errors.values():
                for msg in value: 
                    full_msg += msg + "\n" 
            return JsonResponse({"back_msg": full_msg})

    return render(request, 
                  "app_warehouse/classification_change.html", 
                  {"head_title": "编辑物品分类",
                   "active_navbar": "classification",
                   "form": form})


@login_required(login_url="/")
def classification_delete(request):
    """物品分类删除"""
    class_name = request.POST.get("className", '')
    try:
        classification = models.Classification.objects.get(class_name=class_name)
    except Exception as er:
        logging.error(er)
        return JsonResponse({"back_msg": "数据库出错, 未能正确删除内容."})
    classification.is_deleted = True
    classification.class_name = str(classification.class_name) + "(DELETED %s)"%datetime.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S %f")   # 避免删除后对添加或修改内容时会有唯一性冲突
    classification.save()
    # NOTE: 向history中存入记录
    new_history = History.set_record(cur_user=request.user, model=classification, act=History.DELETE)
    new_history.save()
    return JsonResponse({"back_msg": OK})


@login_required(login_url="/")
def classification_download(request):
    """下载物品分类"""
    class_name = request.GET.get("className", '')
    classifications = models.Classification.objects.filter(is_deleted=False, 
                                                           class_name__icontains=class_name)
    filename = "物品分类 - 搜索名称_%s.xlsx" % class_name if any((class_name,)) else "物品分类.xlsx"
    file_wb = Workbook()
    file_sheet = file_wb.get_sheet_by_name(file_wb.sheetnames[0])
    file_sheet.title = "物品分类"
    file_sheet.append(["物品分类名称", "备注"])
    for info in classifications.values():
        file_sheet.append([info["class_name"], info["remark"].replace('\n', '').replace('\r', '')])
    response = HttpResponse(save_virtual_workbook(file_wb))
    response["Content-Type"] = "application/vnd.ms-excel"
    response["Content-Disposition"] = "attachment;filename=\"%s\"" % filename
    return response


### 销售单管理 ###
